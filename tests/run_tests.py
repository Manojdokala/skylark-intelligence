import unittest
import os
import sys

# Test Normalization & Analytics Logic in Python to verify fixture dataset
import openpyxl

class TestSkylarkIntelligenceBI(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.deals_path = r'd:\7thsem\skylark\Deal funnel Data.xlsx'
        cls.wo_path = r'd:\7thsem\skylark\Work_Order_Tracker Data.xlsx'

    def test_excel_files_exist(self):
        self.assertTrue(os.path.exists(self.deals_path), "Deals excel fixture must exist")
        self.assertTrue(os.path.exists(self.wo_path), "Work Orders excel fixture must exist")

    def test_deals_data_structure(self):
        wb = openpyxl.load_workbook(self.deals_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in list(ws.iter_rows())[0]]
        self.assertIn("Deal Name", headers)
        self.assertIn("Masked Deal value", headers)
        self.assertIn("Sector/service", headers)
        self.assertIn("Close Date (A)", headers)

    def test_work_orders_data_structure(self):
        wb = openpyxl.load_workbook(self.wo_path, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in list(ws.iter_rows())[1]] # Row index 1
        self.assertIn("Deal name masked", headers)
        self.assertIn("Amount in Rupees (Excl of GST) (Masked)", headers)
        self.assertIn("Execution Status", headers)
        self.assertIn("Sector", headers)

    def test_deterministic_pipeline_calculation(self):
        wb = openpyxl.load_workbook(self.deals_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        val_idx = headers.index("Masked Deal value")
        
        total_val = 0
        valid_count = 0
        for row in rows[1:]:
            v = row[val_idx]
            if v is not None:
                try:
                    vf = float(v)
                    total_val += vf
                    valid_count += 1
                except ValueError:
                    pass
        
        self.assertGreater(total_val, 0, "Total pipeline value must be positive")
        self.assertGreater(valid_count, 0, "Valid deals count must be positive")
        print(f"Verified {valid_count} valid deals totaling Rs. {total_val:,.2f}")

    def test_cross_board_sector_overlap(self):
        wb1 = openpyxl.load_workbook(self.deals_path, data_only=True)
        rows1 = list(wb1.active.iter_rows(values_only=True))
        h1 = rows1[0]
        s1_idx = h1.index("Sector/service")
        deals_sectors = set(r[s1_idx] for r in rows1[1:] if r[s1_idx] is not None)

        wb2 = openpyxl.load_workbook(self.wo_path, data_only=True)
        rows2 = list(wb2.active.iter_rows(values_only=True))
        h2 = rows2[1]
        s2_idx = h2.index("Sector")
        wo_sectors = set(r[s2_idx] for r in rows2[1:] if r[s2_idx] is not None)

        overlap = deals_sectors.intersection(wo_sectors)
        self.assertGreater(len(overlap), 0, "There must be overlapping sectors between Deals and Work Orders")
        print(f"Overlapping sectors found across boards: {overlap}")

if __name__ == '__main__':
    unittest.main()
